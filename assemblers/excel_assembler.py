"""Excel Assembler — Step ⑧ for Excel deliverables.

In the Suadeo doc (§5.3.3) this is ClosedXML in C#. Here: openpyxl.
"""

from __future__ import annotations

import uuid
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

OUTPUT_DIR = Path(__file__).parent.parent / "outputs"


def build_excel_workbook(content: dict[str, Any]) -> str:
    """Build a real .xlsx file from a structured content dict.

    Expected shape:
        {
            "title": "...",
            "sheets": [
                {"name": "Summary", "headers": [...], "rows": [[...]]},
                ...
            ]
        }
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    title = content.get("title", "Suadeo_Export")

    for sheet_def in content.get("sheets", []):
        name = sheet_def.get("name", "Sheet")[:31]  # Excel sheet name limit
        ws = wb.create_sheet(name)
        headers = sheet_def.get("headers", [])
        if headers:
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E79")
        for row in sheet_def.get("rows", []):
            ws.append(row)

    safe_title = "".join(c if c.isalnum() else "_" for c in title)[:50]
    filename = f"{safe_title}_{uuid.uuid4().hex[:8]}.xlsx"
    out_path = OUTPUT_DIR / filename
    wb.save(out_path)
    return str(out_path.relative_to(OUTPUT_DIR.parent))
